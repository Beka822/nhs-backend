from fastapi import APIRouter,Depends,Query
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import PieChart,Reference,BarChart
from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from datetime import datetime,timedelta
from sqlalchemy.orm import Session
from sqlalchemy import text
from core.db import get_db
from models.user import User
from core.dependencies import get_user_object
router=APIRouter(prefix="/dashboard",tags=["Dashboard"])
@router.get("/patients")
def get_patient_stats(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    today=db.execute(text("""
                          SELECT COUNT(*) FROM mv_patient_visits
                          WHERE hospital_id=:hospital_id
                          AND visit_date=CURRENT_DATE
                          """),{"hospital_id":
                                hospital_id}).scalar()
    week=db.execute(text("""
                         SELECT COUNT(*) FROM mv_patient_visits
                         WHERE hospital_id=:hospital_id
                         AND visit_date >= CURRENT_DATE -
                         INTERVAL '7 days'
                         """),{"hospital_id":
                               hospital_id}).scalar()
    month=db.execute(text("""
                          SELECT COUNT(*) FROM mv_patient_visits
                          WHERE hospital_id=:hospital_id
                          AND date_trunc('month',visit_date)=
                          date_trunc('month',CURRENT_DATE)
                          """),{"hospital_id":
                                hospital_id}).scalar()
    return {
        "today":today,
        "week":week,
        "month":month
    }
@router.get("/visits-trend")
def get_visits_trend(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    data=db.execute(text("""
                         SELECT visit_date, COUNT(*) as total
                         FROM mv_patient_visits
                         WHERE hospital_id=:hospital_id
                         AND visit_date >= CURRENT_DATE -
                         INTERVAL '30 days'
                         GROUP BY visit_date
                         ORDER BY visit_date
                         """),{"hospital_id":
                               current_user.hospital_id}).fetchall()
    return [
        {"date": row[0], "total": row[1]}
        for row in data
    ]
@router.get("/top-services")
def get_top_services(period:str=Query("month"),db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    period=period.lower()
    now=datetime.utcnow()
    if period == "today":
        start=datetime(now.year,now.month,now.day)
        end=start + timedelta(days=1)
    elif period== "week":
        start=now - timedelta(days=7)
        end=now
    elif period== "month":
        start=datetime(now.year,now.month,1)
        if now.month==12:
            end=datetime(now.year+1,1,1)
        else:
            end=datetime(now.year,now.month+1,1)
    elif period=="year":
        start=datetime(now.year,1,1)
        end=datetime(now.year+1,1,1)
    else:
        start=None
        end=None
    query="""SELECT
    s.name,
    COUNT(*) AS usage_count,
    SUM(bi.total_price) AS total_revenue
    FROM bill_items bi
    JOIN services s ON s.service_id=bi.service_id
    JOIN bills b ON bi.bill_id=b.bill_id
    WHERE b.hospital_id=:hospital_id
    """
    params={"hospital_id":hospital_id}
    if start and end:
        query += "AND b.created_at >= :start AND b.created_at <:end"
        params.update({"start":start,"end":end})
    query += """ GROUP BY s.name
    ORDER BY total_revenue DESC
    LIMIT 5
    """
    data=db.execute(text(query),params).fetchall()
    return [
        {
            "name":row[0],
            "usage_count":row[1],
            "revenue":float(row[2] or 0)
        }
        for row in data
    ]
@router.get("/ward-bor-trend")
def get_ward_bor_trend(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    data=db.execute(text("""
                         SELECT ward_name,date,occupancy_rate
                         FROM mv_ward_bor_trend
                         WHERE hospital_id=:hospital_id
                         ORDER BY ward_name,date
                         """),{"hospital_id":
                               hospital_id}).fetchall()
    return [
        {
            "ward": row[0],
            "date": row[1],
            "occupancy_rate": float(row[2])
        }
        for row in data
    ]
@router.get("/admission-discharge-trend")
def get_admission_discharge_trend(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    data=db.execute(text("""
                         SELECT date,admissions,discharges,net_flow
                         FROM mv_admission_discharge_trend
                         WHERE hospital_id=:hospital_id
                         ORDER BY date
                         """),{"hospital_id":
                               hospital_id}).fetchall()
    return [
        {
            "date": row[0],
            "admissions": row[1],
            "discharges": row[2],
            "net_flow": row[3]
        }
        for row in data
    ]
@router.get("/icu-occupancy")
def get_icu_occupancy(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    data=db.execute(text("""
                         SELECT date,occupied_beds,total_beds,occupancy_rate
                         FROM mv_icu_occupancy
                         WHERE hospital_id=:hospital_id
                         ORDER BY date
                         """),{"hospital_id":
                               hospital_id}).fetchall()
    latest=data[-1] if data else None
    occupancy_rate=float(latest[3]) if latest else 0
    return {
        "trend": [
            {
                "date": row[0],
                "occupied_beds": row[1],
                "total_beds": row[2],
                "occupancy_rate": float(row[3])
            }
            for row in data
        ],
        "current":{
            "occupied_beds":latest[1] if latest else 0,
            "total_beds":latest[2] if latest else 0,
            "occupancy_rate":occupancy_rate
        },
        "alert":{
            "level":(
                "CRITICAL" if occupancy_rate >= 90 
                else
                "WARNING" if occupancy_rate >=75
                else
                "NORMAL"
            ),
            "message":(
                "ICU is critically full" if occupancy_rate >= 90 else
                "ICU nearing capacity" if occupancy_rate >=75 else
                "ICU operating normally"
            )
        }
    }
@router.get("/los-analytics")
def get_los_analytics(period:str="month",db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    now=datetime.utcnow()
    if period =="today":
        start=datetime(now.year,now.month,now.day)
    elif period=="week":
        start=now-timedelta(days=7)
    elif period=="month":
        start=datetime(now.year,now.month,1)
    elif period=="year":
        start=datetime(now.year,1,1)
    else:
        start=None
    query="""SELECT
    ward_name,service_name,
    AVG(avg_los) as avg_los,
    AVG(median_los) as median_los,
    AVG(discharge_efficiency_score) as
    efficiency
    FROM mv_los_analytics
    WHERE hospital_id=:hospital_id
    """
    params={"hospital_id":hospital_id}
    if start:
        query += "AND date >= :start"
        params["start"]=start

    query += """ GROUP BY ward_name,service_name
    ORDER BY efficiency DESC
    LIMIT 10 
    """
    data=db.execute(text(query),params).fetchall()
    return [
        {
            "ward": row[0],
            "service": row[1],
            "avg_los":float(row[2]),
            "median_los":float(row[3]),
            "efficiency_score":float(row[4])
        }
        for row in data
    ]
@router.get("/top-transfer-reasons")
def get_top_transfer_reasons(year:int,month:int,db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    start=datetime(year,month,1)
    data=db.execute(text("""
                         WITH monthly_total AS (
                         SELECT SUM(transfer_count) AS total
                         FROM mv_transfer_reason_analysis
                         WHERE hospital_id=:hospital_id
                         AND month=:start)
                         SELECT
                         m.reason,
                         m.transfer_count,
                         ROUND((m.transfer_count*100.0 / NULLIF(mt.total,0)),2)
                         AS percentage
                         FROM mv_transfer_reason_analysis m
                         CROSS JOIN monthly_total mt
                         WHERE m.hospital_id=:hospital_id
                         AND m.month=:start
                         ORDER BY m.transfer_count DESC
                         LIMIT 3
                         """),{
                             "hospital_id":hospital_id,
                             "start":start
                         }).fetchall()
    return [
        {
            "reason": row[0],
            "count": row[1],
            "percentage":float(row[2] or 0)
        }
        for row in data
    ]
@router.get("/payment-analytics")
def get_payment_analytics(
    period:str="month",
    db:Session=Depends(get_db),
    current_user:User=Depends(get_user_object)
):
    hospital_id=current_user.hospital_id
    today=datetime.utcnow().date()
    if period =="today":
        start=today
        end=today + timedelta(days=1)
    elif period=="week":
        start = today - timedelta(days=today.weekday())
        end=start + timedelta(days=7)
    elif period=="month":
        start=today.replace(day=1)
        end=(start + timedelta(days=32)).replace(day=1)
    elif period=="year":
        start=today.replace(month=1,day=1)
        end=start.replace(year=start.year +1)
    else:
        start=today.replace(day=1)
        end=(start + timedelta(days=32)).replace(day=1)
    distribution=db.execute(text("""
                                 SELECT
                                 payment_method,
                                 SUM(total_amount) AS amount
                                 FROM mv_payment_analytics
                                 WHERE hospital_id=:hospital_id
                                 AND date >=:start AND date <:end
                                 GROUP BY payment_method
                                 """),{
                                     "hospital_id":hospital_id,
                                     "start":start,
                                     "end":end
                                 }).fetchall()
    total=sum(row[1] or 0 for row in distribution)
    pie=[
        {
            "method":row[0],
            "amount": float(row[1] or 0),
            "percentage":round((row[1]/total)*100,2) if total else 0
        }
        for row in distribution
    ]
    trend=db.execute(text("""
                          SELECT
                          date,
                          SUM(total_amount) AS revenue
                          FROM mv_payment_analytics
                          WHERE hospital_id=:hospital_id
                          AND date >=:start AND date<:end
                          GROUP BY date
                          ORDER BY date
                          """),{
                              "hospital_id":hospital_id,
                              "start":start,
                              "end":end
                          }).fetchall()
    trend_data=[
        {
            "date":row[0],
            "revenue":float(row[1] or 0)
        }
        for row in trend
    ]
    digital_cash=db.execute(text("""
                                 SELECT
                                 CASE
                                 WHEN payment_method='Cash' THEN 'Cash'
                                 ELSE 'Digital'
                                 END AS category,
                                 SUM(total_amount) AS amount
                                 FROM mv_payment_analytics
                                 WHERE hospital_id=:hospital_id
                                 AND date >=:start AND date < :end
                                 GROUP BY category
                                 """),{
                                    "hospital_id":hospital_id,
                                    "start":start,
                                    "end":end
                                 }).fetchall()
    digital_cash_data=[
        {
            "category": row[0],
            "amount": float(row[1] or 0)
        }
        for row in digital_cash
    ]
    insurance=db.execute(text("""
                              SELECT
                              SUM(total_amount) FILTER (WHERE
                              payment_method='Insurance') AS
                              insurance_amount,
                              SUM(total_amount) AS total_amount
                              FROM mv_payment_analytics
                              WHERE hospital_id=:hospital_id
                              AND date >=:start AND date < :end
                              """),{
                                  "hospital_id":hospital_id,
                                  "start":start,
                                  "end":end
                              }).fetchone()
    insurance_amount=float(insurance[0] or 0)
    total_amount=float(insurance[1] or 0)
    insurance_dependency=round((insurance_amount/total_amount)*100,2) if total_amount else 0
    return{
        "distribution":pie,
        "trend":trend_data,
        "digital_vs_cash":digital_cash_data,
        "insurance_dependency":insurance_dependency
    }
@router.get("/monthly-excel")
def generate_monthly_report(year:int,month:int,db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    hospital=db.execute(text("""
                             SELECT hospital_name
                             FROM hospitals
                             WHERE hospital_id=:hospital_id
                             """),{
                                 "hospital_id":hospital_id
                             }).fetchone()
    hospital_name=hospital[0] if hospital else "Hospital"
    start=datetime(year,month,1)
    end=(start + timedelta(days=32)).replace(day=1)
    #DATA FETCH
    #visits
    total_visits=db.execute(text("""
                                 SELECT COUNT(*) FROM visits
                                 WHERE hospital_id=:hospital_id
                                 AND created_at >= :start AND created_at <:end
                                 """),{"hospital_id":hospital_id,
                                 "start":start,
                                 "end":end}).scalar()
    #Admissions/Discharges
    admission_data=db.execute(text("""
                                   SELECT
                                   COUNT(*) FILTER (WHERE
                                   admitted_at>=:start AND admitted_at < :end) AS admissions,
                                   COUNT(*) FILTER (WHERE
                                   discharge_at >=:start AND discharge_at <:end) AS discharges
                                   FROM admissions
                                   WHERE hospital_id=:hospital_id
                                   """),{
                                       "hospital_id":hospital_id,
                                       "start":start,
                                       "end":end
                                   }).fetchone()
    admissions=admission_data[0] or 0
    discharges=admission_data[1] or 0
    net_flow=admissions - discharges
    #Revenue
    revenue=db.execute(text("""
                            SELECT SUM(amount) FROM payments
                            WHERE hospital_id=:hospital_id
                            AND received_at >=:start AND
                            received_at <:end
                            """),{
                                "hospital_id":hospital_id,
                                "start":start,
                                "end":end
                            }).scalar() or 0
    #Payment breakdown
    payments=db.execute(text("""
                             SELECT payment_method,SUM(amount)
                             FROM payments
                             WHERE hospital_id=:hospital_id
                             AND received_at >=:start AND received_at <:end
                             GROUP BY payment_method
                             """),{
                                 "hospital_id":hospital_id,
                                 "start":start,
                                 "end":end
                             }).fetchall()
    total_payments=sum([p[1] or 0 for p in payments])
    insurance_amount=next((p[1] for p in payments if p[0]=="Insurance"),0)
    insurance_pct=(insurance_amount/total_payments*100) if total_payments else 0
    #create Pie Chart
    #pie=PieChart()
    #pie.title="Payment Distribution"
    #labels=Reference(ws,min_col=1,min_row=16,max_row=15 + len(payments))
    #data=Reference(ws,min_col=2,min_row=15,max_row=15 + len(payments))
    #pie.add_data(data,titles_from_data=True)
    #pie.set_categories(labels)
    #ws.add_chart(pie,"D8")
    #Top services
    services=db.execute(text("""
                             SELECT name, SUM(revenue)
                             FROM (
                             SELECT s.name,bi.total_price AS revenue
                             FROM bill_items bi
                             JOIN services s ON s.service_id=bi.service_id
                             WHERE bi.hospital_id=:hospital_id)
                             sub
                             GROUP BY name
                             ORDER BY SUM(revenue) DESC
                             LIMIT 5
                             """),{
                                 "hospital_id":hospital_id
                             }).fetchall()
    #bar chart
    #bar=BarChart()
    #bar.title="Top Services Revenue"
    #labels=Reference(ws,min_col=1,min_row=row - len(services),max_row=row -1)
    #data=Reference(ws,min_col=2,min_row=row - len(services) - 1,max_row=row - 1)
    #bar.add_data(data,titles_from_data=True)
    #bar.set_categories(labels)
    #ws.add_chart(bar, "D20")
    #CREATE EXCEL
    wb=Workbook()
    ws=wb.active
    ws.title="Monthly Report"
    #Title
    ws["A1"]=hospital_name.upper()
    ws["A1"].font=Font(size=16,bold=True)
    ws["A2"]="MONTHLY HOSPITAL REPORT"
    ws["A3"]= f"Period: {year}-{month:02d}"
    #PATIENT SECTION
    ws["A4"]="Patient Statistics"
    ws["A4"].font=Font(bold=True)
    ws["A5"]="Total Visits"
    ws["B5"]=total_visits
    #ADMISSION SECTION
    ws["A7"]="Admissions"
    ws["A7"].font=Font(bold=True)
    ws["A8"]="Admissions"
    ws["B8"]=admissions
    ws["A9"]="Discharges"
    ws["B9"]=discharges
    ws["A10"]="Net Flow"
    ws["B10"]=net_flow
    #FINANCIAL SECTION
    ws["A12"]="Financial Summary"
    ws["A12"].font=Font(bold=True)
    ws["A13"]="Total Revenue"
    ws["B13"]=float(revenue)
    #Payment breakdown
    row=15
    ws[f"A{row}"]="Payment Breakdown"
    ws[f"A{row}"].font=Font(bold=True)
    row += 1
    for method,amount in payments:
        ws[f"A{row}"]=method
        ws[f"B{row}"]=float(amount or 0)
        row +=1
    #create Pie Chart
    pie=PieChart()
    pie.title="Payment Distribution"
    labels=Reference(ws,min_col=1,min_row=16,max_row=15 + len(payments))
    data=Reference(ws,min_col=2,min_row=15,max_row=15 + len(payments))
    pie.add_data(data,titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie,"D8")
    #TOP SERVICES
    row +=1
    ws[f"A{row}"]="Top Services"
    ws[f"A{row}"].font=Font(bold=True)
    row +=1
    for name,revenue_val in services:
        ws[f"A{row}"]=name
        ws[f"B{row}"]=float(revenue_val or 0)
        row += 1
     #bar chart
    bar=BarChart()
    bar.title="Top Services Revenue"
    labels=Reference(ws,min_col=1,min_row=row - len(services),max_row=row -1)
    data=Reference(ws,min_col=2,min_row=row - len(services) - 1,max_row=row - 1)
    bar.add_data(data,titles_from_data=True)
    bar.set_categories(labels)
    ws.add_chart(bar, "D20")
    #CHARTS
    ws["D5"]="Insurance Dependency"
    ws["E5"]=f"{insurance_pct:.2f}%"
    ws["D5"].font=Font(bold=True)
    #INSIGHTS
    ws["D2"]="Insights"
    ws["D2"].font=Font(bold=True)
    insight="Balanced revenue streams"
    if insurance_pct > 70:
        insight="High reliance on insurance payments"
    elif insurance_pct < 20:
        insight="Low insurance utilization"
    elif any(p[0]=="Cash" and p[1]> total_payments* 0.5 for p in payments):
        insight="High cash usage - potential fraud risk"
    ws["D3"]=insight
    #AUTO WIDTH
    for col in ["A", "B"]:
        ws.column_dimensions[col].width=25
    #RETURN FILE
    stream=BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f"attachment; filename=report_{year}_{month}.xlsx"
        }
    )
@router.get("/monthly-pdf")
def generate_pdf_report(year:int,month:int,db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    hospital_id=current_user.hospital_id
    hospital=db.execute(text("""
                             SELECT hospital_name FROM hospitals
                             WHERE hospital_id=:hospital_id
                             """),{
                                 "hospital_id":hospital_id
                             }).fetchone()
    hospital_name=hospital[0] if hospital else "Hospital"
    start=datetime(year,month,1)
    end=(start + timedelta(days=32)).replace(day=1)
    visits=db.execute(text("""
                           SELECT COUNT(*) FROM visits
                           WHERE hospital_id=:hospital_id
                           AND created_at >=:start AND created_at < :end
                           """),{
                               "hospital_id":hospital_id,
                               "start":start,
                               "end":end
                           }).scalar()
    revenue=db.execute(text("""
                            SELECT SUM(amount) FROM payments 
                            WHERE hospital_id=:hospital_id
                            AND received_at >=:start AND received_at < :end
                            """),{
                                "hospital_id":hospital_id,
                                "start":start,
                                "end":end
                            }).scalar() or 0
    payments=db.execute(text("""
                             SELECT payment_method,SUM(amount)
                             FROM payments
                             WHERE hospital_id=:hospital_id
                             AND received_at >=:start AND received_at <:end
                             GROUP BY payment_method
                             """),{
                                 "hospital_id":hospital_id,
                                 "start":start,
                                 "end":end
                             }).fetchall()
    total_payments=sum([p[1] or 0 for p in payments])
    insurance_amount=next((p[1] for p in payments if p[0]=="Insurance"),0)
    insurance_pct=(insurance_amount/total_payments*100) if total_payments else 0
    #Admissions and discharges
    admission_data=db.execute(text("""
                                   SELECT
                                   COUNT(*) FILTER (WHERE
                                   admitted_at>=:start AND admitted_at < :end) AS admissions,
                                   COUNT(*) FILTER (WHERE
                                   discharge_at >=:start AND discharge_at <:end) AS discharges
                                   FROM admissions
                                   WHERE hospital_id=:hospital_id
                                   """),{
                                       "hospital_id":hospital_id,
                                       "start":start,
                                       "end":end
                                   }).fetchone()
    admissions=admission_data[0] or 0
    discharges=admission_data[1] or 0
    net_flow=admissions - discharges
    #OCCUPANCY
    occupancy=db.execute(text("""
                              SELECT AVG(occupancy_rate)
                              FROM mv_ward_bor_trend
                              WHERE hospital_id=:hospital_id
                              AND date >=:start AND < :end
                              """),{
                                  "hospital_id":hospital_id,
                                  "start":start,
                                  "end":end
                              }).scalar() or 0
    #LOS
    los=db.execute(text("""
                        SELECT
                        AVG((discharge_at::date - admitted_at::date)) AS avg_los,
                        PERCENTILE_CONT(0.5) WITHIN GROUP(
                        ORDER BY (discharge_at::date - admitted_at::date)) AS median_los
                        FROM admissions
                        WHERE hospital_id=:hospital_id
                        AND discharge_at IS NOT NULL
                        AND discharge_at >= :start AND discharge_at<:end
                        """),{
                            "hospital_id":hospital_id,
                            "start":start,
                            "end":end
                        }).fetchone()
    avg_los=float(los[0] or 0)
    median_los=float(los[1] or 0)
    #DIAGNOSES
    diagnoses=db.execute(text("""
                              SELECT diagnosis,COUNT(*) AS count
                              FROM visits
                              WHERE hospital_id=:hospital_id
                              AND created_at >=:start AND created_at <:end
                              GROUP BY diagnosis
                              ORDER BY count DESC
                              LIMIT 5
                              """),{
                                  "hospital_id":hospital_id,
                                  "start":start,
                                  "end":end
                              }).fetchall()
    #PDF
    buffer=BytesIO()
    doc=SimpleDocTemplate(buffer,pagesize=A4,rightMargin=30,leftMargin=30,topMargin=30,bottomMargin=30)
    styles=getSampleStyleSheet()
    elements=[]
    elements.append(Paragraph("<b>MINISTRY OF HEALTH - KENYA</b>", styles["Normal"]))
    elements.append(Spacer(1,6))
    elements.append(Paragraph(f"<b>{hospital_name}</b>",styles["Title"]))
    elements.append(Spacer(1,6))
    elements.append(Paragraph(f"Monthly Report ({year}-{month:02d})",styles["Heading2"]))
    elements.append(Spacer(1, 12))
    table_data=[
        ["Indicator", "Value"],
        ["Total Visits",visits],
    ]
    table=Table(table_data, colWidths=[250, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0), colors.white),
        ("GRID",(0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1,12))
    elements.append(Paragraph("<b>MOH 711 - INPATIENT SUMMARY</b>",styles["Heading3"]))
    elements.append(Spacer(1,6))
    ipd_table=[
        ["Indicator","Value"],
        ["Admissions",admissions,0],
        ["Discharges",discharges,0],
        ["Net Flow",net_flow,0],
    ]
    table=Table(ipd_table,colWidths=[250,150])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1,12))
    elements.append(Paragraph("<b>BED UTILIZATION</b>",styles["Heading3"]))
    elements.append(Spacer(1,6))
    bed_table=[
        ["Metric", "Value"],
        ["Average Occupancy Rate",occupancy],
    ]
    table=Table(bed_table,colWidths=[250,150])
    table.setStyle([
        ("GRID",(0,0),(-1,-1),1, colors.black),
    ])
    elements.append(table)
    elements.append(Spacer(1,12))
    elements.append(Paragraph("<b>LENGTH OF STAY</b>",styles["Heading3"]))
    elements.append(Spacer(1,6))
    los_table=[
        ["Metric", "Value"],
        ["Average LOS",avg_los],
        ["Median LOS", median_los],
    ]
    table=Table(los_table,colWidths=[250,150])
    table.setStyle([
        ("GRID",(0,0),(-1,-1),1,colors.black),
    ])
    elements.append(table)
    elements.append(Spacer(1,12))
    elements.append(Paragraph("<b>TOP DIAGNOSES</b>",styles["Heading3"]))
    elements.append(Spacer(1,6))
    diag_table=[["Diagnosis","Cases"]]
    for d in diagnoses:
        diag_table.append([d[0],d[1]])
    table=Table(diag_table,colWidths=[250,150])
    table.setStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
    ])
    elements.append(table)
    elements.append(Spacer(1,12))
    elements.append(Paragraph("<b>FINANCIAL SUMMARY</b>",styles["Heading3"]))
    elements.append(Spacer(1,6))
    fin_table=[
        ["Metric", "Value"],
        ["Total_Revenue",revenue],
        ["Insurance Dependency",insurance_pct],
    ]
    table=Table(fin_table,colWidths=[250,150])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1,20))
    insight="Operations stable"
    if occupancy > 85:
        insight="High bed occupancy - capacity strain"
    elif avg_los > 7:
        insight="Long hospital stays - possible inefficiency"
    elif len(diagnoses)>0 and diagnoses[0][1]>50:
        insight=f"High prevalence of {diagnoses[0][0]}"
    elements.append(Paragraph(f"<b>INSIGHTS:</b>{insight}",styles["Normal"]))
    elements.append(Spacer(1,30))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                              styles["Normal"]))
    elements.append(Spacer(1,20))
    elements.append(Paragraph("Prepared by:________________________",
                              styles["Normal"]))
    elements.append(Paragraph("Signature:______________________",
                              styles["Normal"]))
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            f"attachment; filename=report_{year}_{month}.pdf"
        }
    )