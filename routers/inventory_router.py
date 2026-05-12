from io import BytesIO
from fastapi import APIRouter,Depends
from sqlalchemy.orm import Session
from core.db import get_db
from models.user import User
from core.dependencies import get_user_object
from models.pharmacy import Drug
from models.pharmacysale import PharmacySale
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import json
import zipfile
router=APIRouter(prefix="/export",tags=["Export"])
@router.get("/inventory")
def export_inventory(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    drugs=db.query(Drug).filter(Drug.hospital_id==current_user.hospital_id).all()
    wb=Workbook()
    ws=wb.active
    ws.title="Inventory"
    ws.append(["PHARMACY INVENTORY EXPORT"])
    ws["A1"].font=Font(bold=True,size=14)
    ws.append([])
    headers=[
        "Drug ID",
        "Drug Name",
        "Category",
        "Unit",
        "Buying Price",
        "Selling Price",
        "Stock Quantity",
        "Reorder Level",
        "Expiry Date",
        "Created At"
    ]
    ws.append(headers)
    for cell in ws[3]:
        cell.font=Font(bold=True)
    for drug in drugs:
        ws.append([
            drug.drug_id,
            drug.name,
            drug.category,
            drug.unit,
            float(drug.buying_price or 0),
            float(drug.selling_price or 0),
            drug.quantity_in_stock,
            drug.reorder_level,
            str(drug.expiry_date or ""),
            str(drug.created_at)
        ])
    for column in ws.columns:
        max_length=0
        column_letter=column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=max_length + 4
        ws.column_dimensions[column_letter].width=adjusted_width
    stream=BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocuments.spreadsheetml.sheet",
        headers={
            "Content-Disposition":"attachment; filename=inventory_export.xlsx"
        }
    )
@router.get("/sales")
def export_sales(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    sales=db.query(PharmacySale).filter(PharmacySale.hospital_id==current_user.hospital_id).all()
    wb=Workbook()
    ws=wb.active
    ws.title="Pharmacy Sales"
    ws.append(["PHARMACY SALES EXPORT"])
    ws["A1"].font=Font(bold=True,size=14)
    ws.append([])
    headers=[
        "Sale ID",
        "Drug",
        "Quantity",
        "Payment Method",
        "Total Price",
        "Sold At"
    ]
    ws.append(headers)
    for cell in ws[3]:
        cell.font=Font(bold=True)
    for sale in sales:
        drug_name=""
        if sale.drug:
            drug_name=sale.drug.name
        ws.append([
            sale.sale_id,
            drug_name,
            sale.quantity,
            sale.payment_method,
            float(sale.total_price or 0),
            str(sale.sold_at)
        ])
    for column in ws.columns:
        max_length=0
        try:
            column_letter=column[0].column_letter
        except:
            continue
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=max_length + 4
        ws.column_dimensions[column_letter].width=adjusted_width
    stream=BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={
                                 "Content-Disposition":
                                 "attachment;filename=sales_export.xlsx"
                             })
@router.get("/full-backup")
def export_full_backup(db:Session=Depends(get_db),current_user:User=Depends(get_user_object)):
    drugs=db.query(Drug).filter(Drug.hospital_id==current_user.hospital_id).all()
    sales=db.query(PharmacySale).filter(PharmacySale.hospital_id==current_user.hospital_id).all()
    zip_buffer=BytesIO()
    with zipfile.ZipFile(zip_buffer,"w",zipfile.ZIP_DEFLATED) as zip_file:
        inventory_stream=BytesIO()
        wb=Workbook()
        ws=wb.active
        ws.title="Inventory"
        ws.append(["PHARMACY INVENTORY"])
        ws["A1"].font=Font(bold=True,size=14)
        ws.append([])
        headers=[
            "Drug ID",
            "Name",
            "Category",
            "Unit",
            "Buying Price",
            "Selling Price",
            "Stock Quantity",
            "Reorder Level",
            "Expiry Date"
        ]
        ws.append(headers)
        for drug in drugs:
            ws.append([
                drug.drug_id,
                drug.name,
                drug.category,
                drug.unit,
                float(drug.buying_price or 0),
                float(drug.selling_price or 0),
                drug.quantity_in_stock,
                drug.reorder_level,
                str(drug.expiry_date or "")
            ])
        wb.save(inventory_stream)
        inventory_stream.seek(0)
        zip_file.writestr("inventory.xlsx",inventory_stream.read())
        sales_stream=BytesIO()
        wb2=Workbook()
        ws2=wb2.active
        ws2.title="Sales"
        ws2.append(["PHARMACY SALES"])
        ws2["A1"].font=Font(bold=True,size=14)
        ws2.append([])
        headers2=[
            "Sale ID",
            "Drug",
            "Quantity",
            "Payment Method",
            "Total Price",
            "Sold At"
            ]
        ws2.append(headers2)
        for sale in sales:
            drug_name=""
            if sale.drug:
                drug_name=sale.drug.name
            ws2.append([
                sale.sale_id,
                drug_name,
                sale.quantity,
                sale.payment_method,
                float(sale.total_price or 0),
                str(sale.sold_at)
                ])
        wb2.save(sales_stream)
        sales_stream.seek(0)
        zip_file.writestr("sales.xlsx",sales_stream.read())
        summary={
            "hospital_id":current_user.hospital_id,
            "total_drugs":len(drugs),
            "total_sales":len(sales)
            }
        zip_file.writestr("summary.json",json.dumps(summary,indent=2))
    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer,media_type="application/zip",
                             headers={
                                 "Content-Disposition":
                                 "attachment;filename=pharmacy_backup.zip"
                             })
        
